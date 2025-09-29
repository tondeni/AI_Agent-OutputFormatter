# item_definition_rev_xls.py - Item Definition Review Excel formatter
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
from cat.log import log

def create_review_excel(reviews, timestamp):
    """
    Create an Excel file with review results.
    
    Args:
        reviews (list): List of parsed review items
        timestamp (str): Timestamp for file metadata
        
    Returns:
        Workbook: openpyxl Workbook object or None if Excel not available
    """
    if not EXCEL_AVAILABLE:
        log.warning("openpyxl not available - cannot create Excel file")
        return None
    
    log.info(f"Creating Excel review document with {len(reviews)} review items")
    
    wb = openpyxl.Workbook()
    
    # Create main review sheet
    create_review_sheet(wb, reviews)
    
    # Create summary sheet
    create_summary_sheet(wb, reviews)
    
    # Create category breakdown sheet
    create_category_breakdown_sheet(wb, reviews)
    
    log.info("Excel review document created successfully")
    return wb

def create_review_sheet(wb, reviews):
    """
    Create the main review results sheet.
    
    Args:
        wb: openpyxl Workbook object
        reviews (list): List of review items
    """
    ws = wb.active
    ws.title = "Review Results"
    
    # Headers
    headers = ["ID", "Category", "Requirement", "Description", "Status", "Comment", "Hint for Improvement"]
    
    # Create header row with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = create_border()
    
    # Add data rows
    for row, review in enumerate(reviews, 2):
        data = [
            review.get('id', ''),
            review.get('category', ''),
            review.get('requirement', ''),
            review.get('description', ''),
            review.get('status', ''),
            review.get('comment', ''),
            review.get('hint_for_improvement', '')
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = create_border()
            
            # Color code status column
            if col == 5:  # Status column
                apply_status_formatting(cell, value)
    
    # Auto-adjust column widths
    adjust_column_widths(ws, headers, reviews)

def create_summary_sheet(wb, reviews):
    """
    Create summary statistics sheet.
    
    Args:
        wb: openpyxl Workbook object
        reviews (list): List of review items
    """
    summary_ws = wb.create_sheet("Summary")
    
    # Summary statistics
    total_items = len(reviews)
    passed_items = len([r for r in reviews if r.get('status', '').lower() == 'pass'])
    failed_items = len([r for r in reviews if r.get('status', '').lower() == 'fail'])
    partial_items = len([r for r in reviews if 'partial' in r.get('status', '').lower()])
    
    summary_data = [
        ["Review Summary", ""],
        ["Generated on:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ["Total Requirements Reviewed:", total_items],
        ["Passed:", passed_items],
        ["Failed:", failed_items],
        ["Partially Passed:", partial_items],
        ["Compliance Rate:", f"{(passed_items/total_items)*100:.1f}%" if total_items > 0 else "0%"],
        ["", ""],
        ["Status Distribution", "Count"],
        ["Pass", passed_items],
        ["Fail", failed_items],
        ["Partial", partial_items]
    ]
    
    for row, (label, value) in enumerate(summary_data, 1):
        summary_ws.cell(row=row, column=1, value=label)
        summary_ws.cell(row=row, column=2, value=value)
        
        # Style headers
        if label in ["Review Summary", "Status Distribution"]:
            summary_ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        elif label and value:
            summary_ws.cell(row=row, column=1).font = Font(bold=True)
    
    # Auto-adjust summary column widths
    for col in range(1, 3):
        column_letter = get_column_letter(col)
        summary_ws.column_dimensions[column_letter].width = 25

def create_category_breakdown_sheet(wb, reviews):
    """
    Create category breakdown sheet showing results by ISO 26262 category.
    
    Args:
        wb: openpyxl Workbook object
        reviews (list): List of review items
    """
    category_ws = wb.create_sheet("Category Breakdown")
    
    # Group by category
    category_stats = {}
    for review in reviews:
        category = review.get('category', 'General Requirements')
        status = review.get('status', '').lower()
        
        if category not in category_stats:
            category_stats[category] = {'total': 0, 'pass': 0, 'fail': 0, 'partial': 0}
        
        category_stats[category]['total'] += 1
        if 'pass' in status and 'fail' not in status:
            category_stats[category]['pass'] += 1
        elif 'fail' in status:
            category_stats[category]['fail'] += 1
        elif 'partial' in status:
            category_stats[category]['partial'] += 1
    
    # Headers
    headers = ["Category", "Total", "Pass", "Fail", "Partial", "Compliance Rate %"]
    for col, header in enumerate(headers, 1):
        cell = category_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = create_border()
    
    # Data rows
    for row, (category, stats) in enumerate(category_stats.items(), 2):
        compliance_rate = (stats['pass'] / stats['total'] * 100) if stats['total'] > 0 else 0
        
        data = [
            category,
            stats['total'],
            stats['pass'],
            stats['fail'],
            stats['partial'],
            f"{compliance_rate:.1f}%"
        ]
        
        for col, value in enumerate(data, 1):
            cell = category_ws.cell(row=row, column=col, value=value)
            cell.border = create_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Color code compliance rate
            if col == 6:  # Compliance rate column
                if compliance_rate >= 90:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif compliance_rate >= 70:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Auto-adjust column widths
    adjust_column_widths(category_ws, headers, None, min_width=15)

def create_border():
    """Create a standard border style for cells."""
    return Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def apply_status_formatting(cell, status_value):
    """
    Apply color formatting based on status value.
    
    Args:
        cell: openpyxl Cell object
        status_value (str): Status text
    """
    status_lower = str(status_value).lower()
    if 'pass' in status_lower and 'fail' not in status_lower:
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif 'fail' in status_lower:
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif 'partial' in status_lower:
        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

def adjust_column_widths(ws, headers, data=None, min_width=12):
    """
    Auto-adjust column widths based on content.
    
    Args:
        ws: openpyxl Worksheet object
        headers (list): Header row values
        data (list): Data rows (optional)
        min_width (int): Minimum column width
    """
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = len(str(headers[col-1])) if col <= len(headers) else 0
        
        if data:
            # Check data rows for maximum length
            for row in range(2, len(data) + 2):
                try:
                    cell_value = str(ws[f"{column_letter}{row}"].value or "")
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
        
        # Set column width (cap at 50 characters, minimum at min_width)
        adjusted_width = max(min_width, min(max_length + 2, 50))
        ws.column_dimensions[column_letter].width = adjusted_width