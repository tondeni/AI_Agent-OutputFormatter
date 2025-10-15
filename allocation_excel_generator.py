# ==============================================================================
# allocation_excel_generator.py
# Excel file generation for FSR Allocation data
# Place in: AI_Agent-OutputFormatter/generators/allocation_excel_generator.py
# ==============================================================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def generate_allocation_excel(fsrs_data, system_name):
    """
    Generate Excel workbook with FSR Allocation data.
    Per ISO 26262-3:2018, Clause 7.4.2.8
    
    Args:
        fsrs_data: List of FSR dictionaries from working memory
        system_name: Name of the system
    
    Returns:
        openpyxl Workbook object
    """
    
    wb = Workbook()
    
    # Create worksheets
    ws_matrix = wb.active
    ws_matrix.title = "Allocation Matrix"
    
    ws_by_component = wb.create_sheet("By Component")
    ws_by_asil = wb.create_sheet("By ASIL")
    ws_interfaces = wb.create_sheet("Interfaces")
    ws_validation = wb.create_sheet("Validation")
    
    # Populate worksheets
    _create_allocation_matrix_sheet(ws_matrix, fsrs_data, system_name)
    _create_by_component_sheet(ws_by_component, fsrs_data, system_name)
    _create_by_asil_sheet(ws_by_asil, fsrs_data, system_name)
    _create_interfaces_sheet(ws_interfaces, fsrs_data, system_name)
    _create_validation_sheet(ws_validation, fsrs_data, system_name)
    
    return wb


def _create_allocation_matrix_sheet(ws, fsrs_data, system_name):
    """
    Create main allocation matrix showing FSR → Component mapping.
    This is the primary traceability matrix per ISO 26262.
    """
    
    # Title
    ws['A1'] = f"FSR Allocation Matrix - {system_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"ISO 26262-3:2018, Clause 7.4.2.8 - FSR Allocation to Architectural Elements"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:H2')
    
    ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells('A3:H3')
    
    # Headers
    headers = [
        "FSR ID",
        "FSR Description",
        "FSR Type",
        "ASIL",
        "Safety Goal",
        "Allocated To",
        "Component Type",
        "Allocation Rationale"
    ]
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    # Data rows
    row_idx = 6
    for fsr in fsrs_data:
        ws.cell(row=row_idx, column=1).value = fsr.get('id', '')
        ws.cell(row=row_idx, column=2).value = fsr.get('description', '')
        ws.cell(row=row_idx, column=3).value = fsr.get('type', '')
        ws.cell(row=row_idx, column=4).value = fsr.get('asil', '')
        ws.cell(row=row_idx, column=5).value = fsr.get('safety_goal_id', '')
        ws.cell(row=row_idx, column=6).value = fsr.get('allocated_to', 'NOT ALLOCATED')
        ws.cell(row=row_idx, column=7).value = fsr.get('allocation_type', 'Unknown')
        ws.cell(row=row_idx, column=8).value = fsr.get('allocation_rationale', 'N/A')
        
        # Color code by allocation status
        if not fsr.get('allocated_to') or fsr.get('allocated_to') == 'TBD':
            # Red for unallocated
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws.cell(row=row_idx, column=6).fill = fill
        
        # Color code by ASIL
        asil = fsr.get('asil', 'QM')
        asil_cell = ws.cell(row=row_idx, column=4)
        if asil == 'D':
            asil_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif asil == 'C':
            asil_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif asil == 'B':
            asil_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif asil == 'A':
            asil_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Borders
        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="top", wrap_text=True)
        
        row_idx += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 20  # FSR ID
    ws.column_dimensions['B'].width = 50  # Description
    ws.column_dimensions['C'].width = 18  # Type
    ws.column_dimensions['D'].width = 8   # ASIL
    ws.column_dimensions['E'].width = 15  # Safety Goal
    ws.column_dimensions['F'].width = 30  # Allocated To
    ws.column_dimensions['G'].width = 15  # Component Type
    ws.column_dimensions['H'].width = 50  # Rationale
    
    # Freeze panes
    ws.freeze_panes = 'A6'
    
    # Auto-filter
    ws.auto_filter.ref = f"A5:H{row_idx-1}"


def _create_by_component_sheet(ws, fsrs_data, system_name):
    """
    Create sheet showing FSRs grouped by component.
    Shows component-centric view with ASIL considerations.
    """
    
    # Title
    ws['A1'] = f"FSR Allocation by Component - {system_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "Component-centric view showing FSR allocation and ASIL distribution"
    ws.merge_cells('A2:F2')
    
    # Group FSRs by component
    by_component = {}
    unallocated = []
    
    for fsr in fsrs_data:
        component = fsr.get('allocated_to')
        if component and component != 'TBD':
            if component not in by_component:
                by_component[component] = []
            by_component[component].append(fsr)
        else:
            unallocated.append(fsr)
    
    # Headers
    headers = ["Component", "Component Type", "FSR Count", "ASIL Levels", "FSR IDs", "Types"]
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # Data rows
    row_idx = 5
    
    for component, fsrs in sorted(by_component.items()):
        ws.cell(row=row_idx, column=1).value = component
        
        # Component type (get from first FSR)
        comp_type = fsrs[0].get('allocation_type', 'Unknown') if fsrs else 'Unknown'
        ws.cell(row=row_idx, column=2).value = comp_type
        
        # FSR count
        ws.cell(row=row_idx, column=3).value = len(fsrs)
        
        # ASIL levels
        asil_levels = sorted(set(f.get('asil', 'QM') for f in fsrs), reverse=True)
        ws.cell(row=row_idx, column=4).value = ', '.join(asil_levels)
        
        # FSR IDs
        fsr_ids = ', '.join(f.get('id', '') for f in fsrs[:5])
        if len(fsrs) > 5:
            fsr_ids += f" ... (+{len(fsrs) - 5} more)"
        ws.cell(row=row_idx, column=5).value = fsr_ids
        
        # Types
        types = ', '.join(set(f.get('type', '') for f in fsrs))
        ws.cell(row=row_idx, column=6).value = types
        
        # Borders and alignment
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="top", wrap_text=True)
        
        # Highlight high ASIL components
        if 'D' in asil_levels or 'C' in asil_levels:
            ws.cell(row=row_idx, column=4).fill = PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
            )
        
        row_idx += 1
    
    # Unallocated section
    if unallocated:
        row_idx += 1
        cell = ws.cell(row=row_idx, column=1)
        cell.value = "⚠️ UNALLOCATED FSRs"
        cell.font = Font(bold=True, color="FF0000")
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws.merge_cells(f'A{row_idx}:B{row_idx}')
        
        ws.cell(row=row_idx, column=3).value = len(unallocated)
        
        unalloc_ids = ', '.join(f.get('id', '') for f in unallocated[:5])
        if len(unallocated) > 5:
            unalloc_ids += f" ... (+{len(unallocated) - 5} more)"
        ws.cell(row=row_idx, column=5).value = unalloc_ids
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 30
    
    ws.freeze_panes = 'A5'


def _create_by_asil_sheet(ws, fsrs_data, system_name):
    """
    Create sheet showing allocation by ASIL level.
    Critical for verifying ASIL integrity per ISO 26262-3:2018, 7.4.2.8.a
    """
    
    ws['A1'] = f"FSR Allocation by ASIL - {system_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:E1')
    
    ws['A2'] = "ASIL Integrity Check - ISO 26262-3:2018, Clause 7.4.2.8.a"
    ws.merge_cells('A2:E2')
    
    # Group by ASIL
    by_asil = {}
    for fsr in fsrs_data:
        asil = fsr.get('asil', 'QM')
        if asil not in by_asil:
            by_asil[asil] = []
        by_asil[asil].append(fsr)
    
    # Headers
    headers = ["ASIL", "FSR Count", "Allocated", "Unallocated", "Components Used"]
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    row_idx = 5
    for asil in ['D', 'C', 'B', 'A', 'QM']:
        if asil not in by_asil:
            continue
        
        fsrs = by_asil[asil]
        
        ws.cell(row=row_idx, column=1).value = asil
        ws.cell(row=row_idx, column=2).value = len(fsrs)
        
        allocated = [f for f in fsrs if f.get('allocated_to') and f.get('allocated_to') != 'TBD']
        ws.cell(row=row_idx, column=3).value = len(allocated)
        ws.cell(row=row_idx, column=4).value = len(fsrs) - len(allocated)
        
        # Unique components
        components = set(f.get('allocated_to') for f in allocated if f.get('allocated_to'))
        ws.cell(row=row_idx, column=5).value = ', '.join(sorted(components))
        
        # Color code
        if asil == 'D':
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif asil == 'C':
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif asil == 'B':
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif asil == 'A':
            fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        else:
            fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        ws.cell(row=row_idx, column=1).fill = fill
        
        row_idx += 1
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 60


def _create_interfaces_sheet(ws, fsrs_data, system_name):
    """
    Create interfaces sheet showing component interfaces.
    Per ISO 26262-3:2018, 7.4.2.8.c - Interface specifications
    """
    
    ws['A1'] = f"Component Interfaces - {system_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = "Interface Specifications - ISO 26262-3:2018, Clause 7.4.2.8.c"
    ws.merge_cells('A2:D2')
    
    # Headers
    headers = ["FSR ID", "Component", "Interface Specification", "ASIL"]
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows - only FSRs with interface specs
    row_idx = 5
    for fsr in fsrs_data:
        interface = fsr.get('interface', '')
        if interface and interface != 'To be specified in detailed design':
            ws.cell(row=row_idx, column=1).value = fsr.get('id', '')
            ws.cell(row=row_idx, column=2).value = fsr.get('allocated_to', 'N/A')
            ws.cell(row=row_idx, column=3).value = interface
            ws.cell(row=row_idx, column=4).value = fsr.get('asil', '')
            
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="top", wrap_text=True)
            
            row_idx += 1
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 8


def _create_validation_sheet(ws, fsrs_data, system_name):
    """
    Create allocation validation checklist.
    Per ISO 26262-3:2018, 7.4.2.8 requirements.
    """
    
    ws['A1'] = f"Allocation Validation - {system_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = "ISO 26262-3:2018, Clause 7.4.2.8 Compliance Check"
    ws.merge_cells('A2:D2')
    
    # Calculate validation metrics
    total_fsrs = len(fsrs_data)
    allocated = [f for f in fsrs_data if f.get('allocated_to') and f.get('allocated_to') != 'TBD']
    unallocated = total_fsrs - len(allocated)
    
    with_rationale = [f for f in allocated if f.get('allocation_rationale')]
    without_rationale = len(allocated) - len(with_rationale)
    
    # Check ASIL integrity
    asil_issues = []
    for fsr in fsrs_data:
        sg_id = fsr.get('safety_goal_id', '')
        # Note: Can't validate ASIL without safety goal data, flag for manual review
    
    # Headers
    headers = ["Check Item", "Status", "Count", "Compliance"]
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Validation items
    row_idx = 5
    
    checks = [
        ("All FSRs Allocated", len(allocated), total_fsrs, "7.4.2.8"),
        ("Allocation Rationale Provided", len(with_rationale), len(allocated), "7.4.2.8"),
        ("Component Type Specified", 
         len([f for f in allocated if f.get('allocation_type')]), 
         len(allocated), "7.4.2.8"),
    ]
    
    for check_name, actual, expected, clause in checks:
        ws.cell(row=row_idx, column=1).value = check_name
        
        if actual == expected:
            status = "✅ PASS"
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif actual >= expected * 0.8:
            status = "⚠️ PARTIAL"
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            status = "❌ FAIL"
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws.cell(row=row_idx, column=2).value = status
        ws.cell(row=row_idx, column=2).fill = fill
        
        ws.cell(row=row_idx, column=3).value = f"{actual}/{expected}"
        ws.cell(row=row_idx, column=4).value = clause
        
        row_idx += 1
    
    # Summary
    row_idx += 2
    ws.cell(row=row_idx, column=1).value = "Summary:"
    ws.cell(row=row_idx, column=1).font = Font(bold=True)
    
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = f"Total FSRs: {total_fsrs}"
    
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = f"Allocated: {len(allocated)}"
    
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = f"Unallocated: {unallocated}"
    if unallocated > 0:
        ws.cell(row=row_idx, column=1).font = Font(color="FF0000")
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15