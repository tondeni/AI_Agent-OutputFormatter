# ==============================================================================
# allocation_formatter.py
# FSR Allocation Formatter for Excel/Reports
# Formats allocation data per ISO 26262-3:2018, Clause 7.4.2.8
# ==============================================================================

from cat.log import log

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    log.warning("openpyxl not available - Allocation Excel generation disabled")


def create_allocation_excel(fsrs, goals, system_name, timestamp):
    """
    Create Excel workbook with FSR Allocation Matrix.
    
    Per ISO 26262-3:2018, Clause 7.4.2.8:
    - FSRs shall be allocated to architectural elements
    - ASIL shall be inherited (7.4.2.8.a)
    - Freedom from interference considered (7.4.2.8.b)
    - Interface specifications defined (7.4.2.8.c)
    
    Args:
        fsrs: List of FSR dictionaries
        goals: List of Safety Goal dictionaries
        system_name: Name of the system
        timestamp: Timestamp for the document
        
    Returns:
        openpyxl.Workbook or None
    """
    
    if not EXCEL_AVAILABLE:
        log.warning("Cannot create Allocation Excel - openpyxl not available")
        return None
    
    if not fsrs:
        log.warning("No FSRs provided to create Allocation Excel")
        return None
    
    log.info(f"Creating Allocation Excel with {len(fsrs)} FSRs")
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Create sheets
    ws_matrix = wb.active
    ws_matrix.title = "Allocation Matrix"
    
    ws_by_component = wb.create_sheet("By Component")
    ws_by_asil = wb.create_sheet("By ASIL")
    ws_freedom = wb.create_sheet("Freedom From Interference")
    
    # Fill sheets
    create_allocation_matrix_sheet(ws_matrix, fsrs, goals, system_name, timestamp)
    create_by_component_sheet(ws_by_component, fsrs, system_name, timestamp)
    create_by_asil_sheet(ws_by_asil, fsrs, system_name, timestamp)
    create_freedom_interference_sheet(ws_freedom, fsrs, system_name, timestamp)
    
    log.info("✅ Allocation Excel workbook created successfully")
    
    return wb


def create_allocation_matrix_sheet(ws, fsrs, goals, system_name, timestamp):
    """
    Create main allocation matrix sheet.
    Shows complete FSR → Component traceability.
    """
    
    # Title
    ws['A1'] = f"FSR Allocation Matrix - {system_name}"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:I1')
    
    # Metadata
    ws['A2'] = f"Generated: {timestamp}"
    ws['A3'] = "ISO 26262-3:2018, Clause 7.4.2.8 - FSR Allocation to Architectural Elements"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers - YOUR REQUESTED COLUMNS
    headers = [
        "FSR ID",                          # Column 1
        "FSR Description",                 # Column 2
        "FSR Type",                        # Column 3
        "ASIL",                           # Column 4
        "Safety Goal",                    # Column 5
        "Allocated To",                   # Column 6
        "Component Type",                 # Column 7
        "Allocation Rationale",           # Column 8
        "Interface Specification"         # Column 9
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_thin
    
    # Set column widths
    ws.column_dimensions['A'].width = 20  # FSR ID
    ws.column_dimensions['B'].width = 50  # FSR Description
    ws.column_dimensions['C'].width = 18  # FSR Type
    ws.column_dimensions['D'].width = 8   # ASIL
    ws.column_dimensions['E'].width = 15  # Safety Goal
    ws.column_dimensions['F'].width = 30  # Allocated To
    ws.column_dimensions['G'].width = 15  # Component Type
    ws.column_dimensions['H'].width = 50  # Rationale
    ws.column_dimensions['I'].width = 40  # Interface
    
    # Freeze header row
    ws.freeze_panes = 'A6'
    
    # Data rows
    row_idx = 6
    for fsr in fsrs:
        fsr_id = fsr.get('id', 'Unknown')
        description = fsr.get('description', 'N/A')
        fsr_type = fsr.get('type', 'N/A')
        asil = fsr.get('asil', 'N/A')
        sg_id = fsr.get('safety_goal_id', 'N/A')
        allocated_to = fsr.get('allocated_to', 'NOT ALLOCATED')
        comp_type = fsr.get('allocation_type', 'Unknown')
        rationale = fsr.get('allocation_rationale', 'N/A')
        interface = fsr.get('interface', 'To be specified in detailed design')
        
        # Write data in YOUR requested order
        ws.cell(row=row_idx, column=1).value = fsr_id
        ws.cell(row=row_idx, column=2).value = description
        ws.cell(row=row_idx, column=3).value = fsr_type
        ws.cell(row=row_idx, column=4).value = asil
        ws.cell(row=row_idx, column=5).value = sg_id
        ws.cell(row=row_idx, column=6).value = allocated_to
        ws.cell(row=row_idx, column=7).value = comp_type
        ws.cell(row=row_idx, column=8).value = rationale
        ws.cell(row=row_idx, column=9).value = interface
        
        # Style data cells
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border_thin
        
        # Color code allocation status (column 6)
        alloc_cell = ws.cell(row=row_idx, column=6)
        if not allocated_to or allocated_to in ['NOT ALLOCATED', 'TBD', 'N/A']:
            alloc_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            alloc_cell.font = Font(color="9C0006", bold=True)
        
        # Color code by ASIL (column 4)
        asil_cell = ws.cell(row=row_idx, column=4)
        if asil == 'D':
            asil_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif asil == 'C':
            asil_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif asil == 'B':
            asil_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif asil == 'A':
            asil_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        row_idx += 1
    
    # Auto-filter
    ws.auto_filter.ref = f"A5:I{row_idx - 1}"
    
    log.info(f"✅ Created allocation matrix with {row_idx - 6} FSRs")


def create_by_component_sheet(ws, fsrs, system_name, timestamp):
    """
    Create component-centric allocation view.
    Groups FSRs by component to show workload distribution.
    """
    
    # Title
    ws['A1'] = f"Allocation by Component - {system_name}"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:G1')
    
    ws['A2'] = f"Generated: {timestamp}"
    ws['A3'] = "Component-centric view showing FSR allocation distribution"
    
    # Group FSRs by component
    by_component = {}
    unallocated = []
    
    for fsr in fsrs:
        component = fsr.get('allocated_to')
        if component and component not in ['TBD', 'NOT ALLOCATED', 'N/A']:
            if component not in by_component:
                by_component[component] = []
            by_component[component].append(fsr)
        else:
            unallocated.append(fsr)
    
    # Headers
    headers = [
        "Component Name",
        "Component Type",
        "FSR Count",
        "ASIL Levels",
        "Highest ASIL",
        "FSR Types",
        "FSR IDs"
    ]
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    row_idx = 6
    
    for component, comp_fsrs in sorted(by_component.items()):
        ws.cell(row=row_idx, column=1).value = component
        
        # Component type
        comp_type = comp_fsrs[0].get('allocation_type', 'Unknown') if comp_fsrs else 'Unknown'
        ws.cell(row=row_idx, column=2).value = comp_type
        
        # FSR count
        ws.cell(row=row_idx, column=3).value = len(comp_fsrs)
        
        # ASIL levels
        asil_levels = sorted(set(f.get('asil', 'QM') for f in comp_fsrs), reverse=True)
        ws.cell(row=row_idx, column=4).value = ', '.join(asil_levels)
        
        # Highest ASIL
        highest_asil = asil_levels[0] if asil_levels else 'QM'
        ws.cell(row=row_idx, column=5).value = highest_asil
        
        # Color code by highest ASIL
        asil_cell = ws.cell(row=row_idx, column=5)
        if highest_asil == 'D':
            asil_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif highest_asil == 'C':
            asil_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif highest_asil == 'B':
            asil_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # FSR types
        types = ', '.join(sorted(set(f.get('type', '') for f in comp_fsrs)))
        ws.cell(row=row_idx, column=6).value = types
        
        # FSR IDs
        fsr_ids = ', '.join(f.get('id', '') for f in comp_fsrs[:5])
        if len(comp_fsrs) > 5:
            fsr_ids += f" ... (+{len(comp_fsrs) - 5} more)"
        ws.cell(row=row_idx, column=7).value = fsr_ids
        
        row_idx += 1
    
    # Unallocated section
    if unallocated:
        row_idx += 1
        cell = ws.cell(row=row_idx, column=1)
        cell.value = "⚠️ UNALLOCATED FSRs"
        cell.font = Font(bold=True, color="9C0006")
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws.merge_cells(f'A{row_idx}:B{row_idx}')
        
        ws.cell(row=row_idx, column=3).value = len(unallocated)
        
        unalloc_ids = ', '.join(f.get('id', '') for f in unallocated[:10])
        if len(unallocated) > 10:
            unalloc_ids += f" ... (+{len(unallocated) - 10} more)"
        ws.cell(row=row_idx, column=7).value = unalloc_ids
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 50
    
    ws.freeze_panes = 'A6'


def create_by_asil_sheet(ws, fsrs, system_name, timestamp):
    """
    Create ASIL-centric allocation view.
    Critical for verifying ASIL integrity per ISO 26262-3:2018, 7.4.2.8.a
    """
    
    ws['A1'] = f"Allocation by ASIL - {system_name}"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Generated: {timestamp}"
    ws['A3'] = "ASIL Integrity Verification - ISO 26262-3:2018, Clause 7.4.2.8.a"
    
    # Group by ASIL
    by_asil = {}
    for fsr in fsrs:
        asil = fsr.get('asil', 'QM')
        if asil not in by_asil:
            by_asil[asil] = []
        by_asil[asil].append(fsr)
    
    # Headers
    headers = [
        "ASIL Level",
        "Total FSRs",
        "Allocated",
        "Unallocated",
        "Components Used",
        "Component Types"
    ]
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    row_idx = 6
    for asil in ['D', 'C', 'B', 'A', 'QM']:
        if asil not in by_asil:
            continue
        
        asil_fsrs = by_asil[asil]
        
        ws.cell(row=row_idx, column=1).value = asil
        ws.cell(row=row_idx, column=2).value = len(asil_fsrs)
        
        allocated = [f for f in asil_fsrs if f.get('allocated_to') and f.get('allocated_to') not in ['TBD', 'NOT ALLOCATED', 'N/A']]
        ws.cell(row=row_idx, column=3).value = len(allocated)
        ws.cell(row=row_idx, column=4).value = len(asil_fsrs) - len(allocated)
        
        # Components
        components = set(f.get('allocated_to') for f in allocated)
        components.discard(None)
        ws.cell(row=row_idx, column=5).value = ', '.join(sorted(components))
        
        # Component types
        comp_types = set(f.get('allocation_type') for f in allocated)
        comp_types.discard(None)
        comp_types.discard('Unknown')
        ws.cell(row=row_idx, column=6).value = ', '.join(sorted(comp_types))
        
        # Color code ASIL column
        asil_cell = ws.cell(row=row_idx, column=1)
        if asil == 'D':
            asil_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif asil == 'C':
            asil_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif asil == 'B':
            asil_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif asil == 'A':
            asil_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Highlight unallocated
        if len(asil_fsrs) - len(allocated) > 0:
            ws.cell(row=row_idx, column=4).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        row_idx += 1
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 30


def create_freedom_interference_sheet(ws, fsrs, system_name, timestamp):
    """
    Create Freedom from Interference analysis sheet.
    Per ISO 26262-3:2018, 7.4.2.8.b
    
    Checks for potential interference issues:
    - Mixed ASIL levels in same component
    - Resource sharing concerns
    """
    
    ws['A1'] = f"Freedom from Interference Analysis - {system_name}"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Generated: {timestamp}"
    ws['A3'] = "ISO 26262-3:2018, Clause 7.4.2.8.b - Freedom from Interference Considerations"
    
    # Analyze components for ASIL mixing
    component_analysis = {}
    
    for fsr in fsrs:
        component = fsr.get('allocated_to')
        if not component or component in ['TBD', 'NOT ALLOCATED', 'N/A']:
            continue
        
        if component not in component_analysis:
            component_analysis[component] = {
                'asil_levels': set(),
                'fsrs': [],
                'comp_type': fsr.get('allocation_type', 'Unknown')
            }
        
        component_analysis[component]['asil_levels'].add(fsr.get('asil', 'QM'))
        component_analysis[component]['fsrs'].append(fsr)
    
    # Headers
    headers = [
        "Component",
        "Component Type",
        "ASIL Levels",
        "Risk Level",
        "FSR Count",
        "Interference Considerations"
    ]
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    
    # Data rows
    row_idx = 6
    
    for component, analysis in sorted(component_analysis.items()):
        ws.cell(row=row_idx, column=1).value = component
        ws.cell(row=row_idx, column=2).value = analysis['comp_type']
        
        asil_levels = sorted(analysis['asil_levels'], reverse=True)
        ws.cell(row=row_idx, column=3).value = ', '.join(asil_levels)
        
        # Determine risk level
        if len(asil_levels) > 1:
            if 'D' in asil_levels or 'C' in asil_levels:
                risk = "HIGH"
                risk_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                considerations = "Multiple ASIL levels including C/D - Requires spatial/temporal independence, partitioning, and protection mechanisms"
            else:
                risk = "MEDIUM"
                risk_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                considerations = "Multiple ASIL levels - Consider resource partitioning and independence measures"
        else:
            risk = "LOW"
            risk_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            considerations = "Single ASIL level - Standard protection measures apply"
        
        ws.cell(row=row_idx, column=4).value = risk
        ws.cell(row=row_idx, column=4).fill = risk_fill
        ws.cell(row=row_idx, column=4).font = Font(bold=True)
        
        ws.cell(row=row_idx, column=5).value = len(analysis['fsrs'])
        ws.cell(row=row_idx, column=6).value = considerations
        
        # Alignment
        for col_idx in range(1, 7):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="top", wrap_text=True)
        
        row_idx += 1
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 60
    
    ws.freeze_panes = 'A6'
    
    log.info(f"✅ Created freedom from interference analysis for {len(component_analysis)} components")