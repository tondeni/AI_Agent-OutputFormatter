# fsr_formatter_xls.py - IMPROVED VERSION
# Functional Safety Requirements Excel Formatter
# Generates Excel file with FSRs per ISO 26262-3:2018, Clause 7.4.2

from cat.log import log
import re

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    log.warning("openpyxl not available - FSR Excel generation disabled")


def create_fsr_excel(fsrs, system_name, timestamp):
    """
    Create Excel workbook with Functional Safety Requirements.
    
    Args:
        fsrs: List of FSR dictionaries
        system_name: Name of the system
        timestamp: Timestamp for the document
        
    Returns:
        openpyxl.Workbook or None
    """
    
    if not EXCEL_AVAILABLE:
        log.warning("Cannot create FSR Excel - openpyxl not available")
        return None
    
    if not fsrs:
        log.warning("No FSRs provided to create Excel")
        return None
    
    log.info(f"Creating FSR Excel with {len(fsrs)} requirements")
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Create sheets
    ws_summary = wb.active
    ws_summary.title = "FSR Summary"
    
    ws_details = wb.create_sheet("FSR Details")
    
    # Fill Summary Sheet
    create_summary_sheet(ws_summary, fsrs, system_name, timestamp)
    
    # Fill Details Sheet with YOUR requested columns
    create_details_sheet(ws_details, fsrs, system_name, timestamp)
    
    log.info("✅ FSR Excel workbook created successfully")
    
    return wb


def create_summary_sheet(ws, fsrs, system_name, timestamp):
    """Create summary sheet with FSR overview."""
    
    # Title
    ws['A1'] = f"Functional Safety Requirements - {system_name}"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:G1')
    
    # Metadata
    ws['A2'] = f"Generated: {timestamp}"
    ws['A3'] = f"Total FSRs: {len(fsrs)}"
    ws['A4'] = "ISO 26262-3:2018, Clause 7.4.2"
    
    # Statistics
    ws['A6'] = "FSR Statistics by ASIL:"
    ws['A6'].font = Font(bold=True)
    
    asil_counts = {}
    for fsr in fsrs:
        asil = fsr.get('asil', 'Unknown')
        asil_counts[asil] = asil_counts.get(asil, 0) + 1
    
    row = 7
    for asil in ['D', 'C', 'B', 'A', 'QM']:
        if asil in asil_counts:
            ws[f'A{row}'] = f"ASIL {asil}:"
            ws[f'B{row}'] = asil_counts[asil]
            row += 1
    
    # Type statistics
    ws[f'A{row + 1}'] = "FSR Statistics by Type:"
    ws[f'A{row + 1}'].font = Font(bold=True)
    
    type_counts = {}
    for fsr in fsrs:
        ftype = fsr.get('type', 'Unknown')
        type_counts[ftype] = type_counts.get(ftype, 0) + 1
    
    row += 2
    for ftype, count in sorted(type_counts.items()):
        ws[f'A{row}'] = f"{ftype}:"
        ws[f'B{row}'] = count
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15


def create_details_sheet(ws, fsrs, system_name, timestamp):
    """
    Create details sheet with all FSR information.
    
    YOUR REQUESTED COLUMNS:
    1) FSR ID
    2) FSR Description
    3) FSR Allocation
    4) FSR ASIL
    5) Linked Safety Goal
    6) Validation and verification criteria
    7) Time constraints FTTI (if any)
    """
    
    # Define header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers - YOUR REQUESTED FORMAT
    headers = [
        "FSR ID",                                    # Column 1
        "FSR Description",                           # Column 2
        "FSR Allocation",                            # Column 3
        "FSR ASIL",                                  # Column 4
        "Linked Safety Goal",                        # Column 5
        "Validation and Verification Criteria",      # Column 6
        "Time Constraints FTTI (if any)"            # Column 7
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_thin
    
    # Set column widths
    ws.column_dimensions['A'].width = 20  # FSR ID
    ws.column_dimensions['B'].width = 60  # FSR Description
    ws.column_dimensions['C'].width = 25  # FSR Allocation
    ws.column_dimensions['D'].width = 8   # FSR ASIL
    ws.column_dimensions['E'].width = 20  # Linked Safety Goal
    ws.column_dimensions['F'].width = 40  # Validation and Verification Criteria
    ws.column_dimensions['G'].width = 15  # Time Constraints FTTI
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Data rows
    row_idx = 2
    for fsr in fsrs:
        fsr_id = fsr.get('id', 'Unknown')
        description = fsr.get('description', 'N/A')
        allocation = fsr.get('allocated_to', 'N/A')
        asil = fsr.get('asil', 'N/A')
        linked_sg = fsr.get('safety_goal_id', 'N/A')
        verification = fsr.get('verification_criteria', 'N/A')
        ftti = fsr.get('ftti', 'N/A')  # NEW: FTTI column
        
        # Write data in YOUR requested order
        ws.cell(row=row_idx, column=1).value = fsr_id
        ws.cell(row=row_idx, column=2).value = description
        ws.cell(row=row_idx, column=3).value = allocation
        ws.cell(row=row_idx, column=4).value = asil
        ws.cell(row=row_idx, column=5).value = linked_sg
        ws.cell(row=row_idx, column=6).value = verification
        ws.cell(row=row_idx, column=7).value = ftti
        
        # Style data cells
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border_thin
            
            # Color code by ASIL (column 4)
            if col_idx == 4:  # ASIL column
                if asil == 'D':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif asil == 'C':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif asil == 'B':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif asil == 'A':
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        row_idx += 1
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:G{row_idx - 1}"
    
    log.info(f"✅ Created FSR details sheet with {row_idx - 1} requirements")


def parse_fsrs(llm_response, safety_goals):
    """
    IMPROVED: Parse FSRs from LLM response.
    Handles both markdown format AND table format.
    
    Supports:
    1. Markdown format with **Field:** value
    2. Table format with | Field | Value |
    3. Multi-line descriptions
    """
    fsrs = []
    current_sg = None
    current_fsr = None
    current_field = None
    
    lines = llm_response.split('\n')
    
    type_mapping = {
        'AVD': 'Fault Avoidance',
        'DET': 'Fault Detection',
        'CTL': 'Fault Control',
        'SST': 'Safe State Transition',
        'TOL': 'Fault Tolerance',
        'WRN': 'Warning/Indication',
        'TIM': 'Timing',
        'ARB': 'Arbitration'
    }
    
    log.info("🔍 Starting IMPROVED FSR parsing...")
    
    # First, try to detect if this is a table format
    is_table_format = any('|' in line and 'FSR-' in line for line in lines[:50])
    
    if is_table_format:
        log.info("📊 Detected TABLE format - using table parser")
        return parse_fsrs_table_format(llm_response, safety_goals)
    else:
        log.info("📝 Detected MARKDOWN format - using markdown parser")
        return parse_fsrs_markdown_format(llm_response, safety_goals, type_mapping)


def parse_fsrs_table_format(llm_response, safety_goals):
    """
    Parse FSRs from table format:
    | FSR ID | Description | Allocation | ASIL | Linked SG | Verification | FTTI |
    """
    fsrs = []
    lines = llm_response.split('\n')
    
    header_found = False
    header_indices = {}
    
    for line in lines:
        line_stripped = line.strip()
        
        # Skip separator lines
        if not line_stripped or line_stripped.startswith('|---') or line_stripped.startswith('|-'):
            continue
        
        # Check if this is a table row
        if '|' not in line_stripped:
            continue
        
        cells = [cell.strip() for cell in line_stripped.split('|')]
        cells = [c for c in cells if c]  # Remove empty cells
        
        if not cells:
            continue
        
        # Detect header row
        if not header_found:
            if any('fsr' in c.lower() and 'id' in c.lower() for c in cells):
                header_found = True
                # Map header positions
                for idx, cell in enumerate(cells):
                    cell_lower = cell.lower()
                    if 'fsr' in cell_lower and 'id' in cell_lower:
                        header_indices['id'] = idx
                    elif 'description' in cell_lower:
                        header_indices['description'] = idx
                    elif 'allocation' in cell_lower:
                        header_indices['allocation'] = idx
                    elif 'asil' in cell_lower and 'linked' not in cell_lower:
                        header_indices['asil'] = idx
                    elif 'linked' in cell_lower or 'safety goal' in cell_lower:
                        header_indices['linked_sg'] = idx
                    elif 'verification' in cell_lower or 'validation' in cell_lower:
                        header_indices['verification'] = idx
                    elif 'ftti' in cell_lower or 'time' in cell_lower:
                        header_indices['ftti'] = idx
                log.info(f"📋 Found table headers: {header_indices}")
                continue
        
        # Parse data rows
        if header_found and 'FSR-' in cells[0]:
            fsr = {
                'id': cells[header_indices.get('id', 0)] if header_indices.get('id', 0) < len(cells) else 'Unknown',
                'description': cells[header_indices.get('description', 1)] if header_indices.get('description', 1) < len(cells) else 'N/A',
                'allocated_to': cells[header_indices.get('allocation', 2)] if header_indices.get('allocation', 2) < len(cells) else 'N/A',
                'asil': cells[header_indices.get('asil', 3)] if header_indices.get('asil', 3) < len(cells) else 'N/A',
                'safety_goal_id': cells[header_indices.get('linked_sg', 4)] if header_indices.get('linked_sg', 4) < len(cells) else 'N/A',
                'verification_criteria': cells[header_indices.get('verification', 5)] if header_indices.get('verification', 5) < len(cells) else 'N/A',
                'ftti': cells[header_indices.get('ftti', 6)] if header_indices.get('ftti', 6) < len(cells) else 'N/A',
                'type': 'General',
                'operating_modes': 'All modes',
                'safety_goal': 'See Linked SG',
                'safe_state': '',
                'emergency_operation': '',
                'functional_redundancy': ''
            }
            
            # Determine FSR type from ID
            for type_code in ['AVD', 'DET', 'CTL', 'SST', 'TOL', 'WRN', 'TIM', 'ARB']:
                if f'-{type_code}-' in fsr['id']:
                    fsr['type'] = type_code
                    break
            
            fsrs.append(fsr)
            log.debug(f"✅ Parsed FSR from table: {fsr['id']}")
    
    log.info(f"📊 Parsed {len(fsrs)} FSRs from table format")
    return fsrs


def parse_fsrs_markdown_format(llm_response, safety_goals, type_mapping):
    """
    IMPROVED: Parse FSRs from markdown format with better multi-line handling.
    """
    fsrs = []
    current_sg = None
    current_fsr = None
    current_field = None
    accumulating_description = False
    
    lines = llm_response.split('\n')
    
    log.info("🔍 Starting markdown FSR parsing...")
    
    for idx, line in enumerate(lines):
        line_stripped = line.strip()
        
        # Detect safety goal section
        if '## FSRs for Safety Goal:' in line_stripped or 'FSRs for Safety Goal:' in line_stripped:
            for sg in safety_goals:
                if sg['id'] in line_stripped:
                    current_sg = sg
                    log.info(f"📍 Found section for {current_sg['id']}")
                    break
        
        # Detect new FSR ID
        if current_sg and ('FSR-' in line_stripped and ('**FSR-' in line_stripped or line_stripped.startswith('FSR-'))):
            # Save previous FSR
            if current_fsr:
                fsrs.append(current_fsr)
            
            # Extract FSR ID
            fsr_id = line_stripped.replace('**', '').replace('*', '').strip()
            if ' ' in fsr_id:
                fsr_id = fsr_id.split()[0]
            
            # Ensure proper SG prefix
            if not fsr_id.startswith('FSR-SG-'):
                match = re.search(r'FSR-(\d+)', fsr_id)
                if match and current_sg:
                    sg_num = match.group(1)
                    fsr_id = fsr_id.replace(f'FSR-{sg_num}', f'FSR-{current_sg["id"]}')
            
            # Determine type
            fsr_type = 'General'
            for type_code, type_name in type_mapping.items():
                if f'-{type_code}-' in fsr_id:
                    fsr_type = type_name
                    break
            
            current_fsr = {
                'id': fsr_id,
                'safety_goal_id': current_sg['id'],
                'safety_goal': current_sg['description'],
                'asil': current_sg['asil'],
                'type': fsr_type,
                'description': '',
                'operating_modes': '',
                'allocated_to': '',
                'verification_criteria': '',
                'ftti': current_sg.get('ftti', ''),
                'safe_state': current_sg.get('safe_state', ''),
                'emergency_operation': '',
                'functional_redundancy': ''
            }
            accumulating_description = False
            current_field = None
            log.debug(f"🆕 Created FSR: {fsr_id}")
            continue
        
        # Extract FSR fields
        if current_fsr and line_stripped:
            # Remove leading markers
            clean_line = line_stripped.lstrip('- ').lstrip('* ').strip()
            
            # Check for field pattern: **Field:** value or - **Field:** value
            if '**' in clean_line and ':' in clean_line:
                parts = clean_line.split('**')
                if len(parts) >= 2:
                    field_and_value = parts[1]
                    if ':' in field_and_value:
                        field_name = field_and_value.split(':')[0].strip().lower()
                        field_value = ':'.join(field_and_value.split(':')[1:]).strip()
                        
                        # Reset accumulation flag when new field starts
                        accumulating_description = False
                        current_field = field_name
                        
                        # Map fields
                        if 'description' in field_name:
                            current_fsr['description'] = field_value
                            accumulating_description = True
                            log.debug(f"  📝 Description start: {field_value[:50]}...")
                        
                        elif 'asil' in field_name and 'linked' not in field_name:
                            current_fsr['asil'] = field_value
                            log.debug(f"  🏷️ ASIL: {field_value}")
                        
                        elif 'operating mode' in field_name:
                            current_fsr['operating_modes'] = field_value
                            log.debug(f"  ⚙️ Modes: {field_value}")
                        
                        elif 'allocation' in field_name or 'allocated' in field_name:
                            current_fsr['allocated_to'] = field_value
                            log.debug(f"  📍 Allocation: {field_value}")
                        
                        elif 'verification' in field_name or 'validation' in field_name:
                            current_fsr['verification_criteria'] = field_value
                            log.debug(f"  ✓ Verification: {field_value[:50]}...")
                        
                        elif 'ftti' in field_name or 'time constraint' in field_name:
                            current_fsr['ftti'] = field_value
                            log.debug(f"  ⏱️ FTTI: {field_value}")
            
            # Handle multi-line descriptions (continuation lines)
            elif accumulating_description and clean_line and not clean_line.startswith('#'):
                # This is a continuation of the description
                if current_fsr['description']:
                    current_fsr['description'] += ' ' + clean_line
                    log.debug(f"  📝 Description continued: {clean_line[:50]}...")
    
    # Save last FSR
    if current_fsr:
        fsrs.append(current_fsr)
    
    log.info(f"✅ Parsed {len(fsrs)} FSRs from markdown format")
    
    # Debug: show sample
    if fsrs:
        log.info(f"📊 Sample FSR: {fsrs[0]['id']}")
        log.info(f"  Description: {fsrs[0]['description'][:80] if fsrs[0]['description'] else 'EMPTY!'}")
    
    return fsrs