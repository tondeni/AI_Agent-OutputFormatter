# hara_rev_xls.py - HARA Review Excel formatter
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
from cat.log import log


def create_hara_review_excel(reviews, timestamp):
    """
    Create an Excel file with HARA review results.
    
    Args:
        reviews (list): List of parsed review items
        timestamp (str): Timestamp for file metadata
        
    Returns:
        Workbook: openpyxl Workbook object or None if Excel not available
    """
    if not EXCEL_AVAILABLE:
        log.warning("openpyxl not available - cannot create Excel file")
        return None
    
    log.info(f"Creating Excel HARA review document with {len(reviews)} review items")
    
    wb = openpyxl.Workbook()
    
    # Create main review sheet
    create_review_sheet(wb, reviews)
    
    # Create summary sheet
    create_summary_sheet(wb, reviews)
    
    # Create category breakdown sheet
    create_category_breakdown_sheet(wb, reviews)
    
    log.info("Excel HARA review document created successfully")
    return wb


def create_review_sheet(wb, reviews):
    """Create the main review results sheet."""
    ws = wb.active
    ws.title = "HARA Review Results"
    
    # Define headers
    headers = [
        "ID", "Category", "Requirement", "Description", 
        "Status", "Comment", "Hint for Improvement"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="00467F", end_color="00467F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Write review data
    for row_idx, review in enumerate(reviews, 2):
        data = [
            review.get('id', 'N/A'),
            review.get('category', 'N/A'),
            review.get('requirement', 'N/A'),
            review.get('description', 'N/A'),
            review.get('status', 'N/A'),
            review.get('comment', 'N/A'),
            review.get('hint_for_improvement', 'N/A')
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Color code status column
            if col_idx == 5:  # Status column
                format_status_cell(cell, value)
    
    # Adjust column widths
    column_widths = {
        'A': 15,  # ID
        'B': 25,  # Category
        'C': 35,  # Requirement
        'D': 45,  # Description
        'E': 15,  # Status
        'F': 50,  # Comment
        'G': 50   # Hint
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add auto-filter
    ws.auto_filter.ref = f"A1:G{len(reviews) + 1}"


def format_status_cell(cell, status):
    """Format status cell with appropriate colors."""
    status_lower = status.lower()
    
    if 'pass' in status_lower and 'partial' not in status_lower:
        cell.font = Font(bold=True, color="006100")
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif 'fail' in status_lower and 'partial' not in status_lower:
        cell.font = Font(bold=True, color="9C0006")
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif 'partial' in status_lower:
        cell.font = Font(bold=True, color="9C5700")
        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    elif 'not applicable' in status_lower or 'n/a' in status_lower:
        cell.font = Font(color="7F7F7F")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def create_summary_sheet(wb, reviews):
    """Create summary statistics sheet."""
    ws = wb.create_sheet("Summary")
    
    # Calculate statistics
    total = len(reviews)
    pass_count = sum(1 for r in reviews if 'pass' in r.get('status', '').lower() and 'partial' not in r.get('status', '').lower())
    fail_count = sum(1 for r in reviews if 'fail' in r.get('status', '').lower() and 'partial' not in r.get('status', '').lower())
    partial_count = sum(1 for r in reviews if 'partial' in r.get('status', '').lower())
    na_count = sum(1 for r in reviews if 'not applicable' in r.get('status', '').lower() or 'n/a' in r.get('status', '').lower())
    
    applicable = total - na_count
    compliance_rate = (pass_count / applicable * 100) if applicable > 0 else 0
    
    # Title
    ws['A1'] = "ISO 26262-3 HARA Review Summary"
    ws['A1'].font = Font(bold=True, size=16, color="00467F")
    ws.merge_cells('A1:B1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, color="7F7F7F")
    ws.merge_cells('A2:B2')
    
    # Statistics table
    stats = [
        ("", ""),
        ("Metric", "Value"),
        ("Total Review Items", total),
        ("✅ Pass", pass_count),
        ("❌ Fail", fail_count),
        ("⚠️ Partial Pass", partial_count),
        ("➖ Not Applicable", na_count),
        ("", ""),
        ("Applicable Items", applicable),
        ("Compliance Rate", f"{compliance_rate:.1f}%")
    ]
    
    for row_idx, (label, value) in enumerate(stats, 1):
        ws.cell(row=row_idx, column=1).value = label
        ws.cell(row=row_idx, column=2).value = value
        
        # Format header row
        if row_idx == 2:
            ws.cell(row=row_idx, column=1).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row_idx, column=2).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="00467F", end_color="00467F", fill_type="solid")
            ws.cell(row=row_idx, column=2).fill = PatternFill(start_color="00467F", end_color="00467F", fill_type="solid")
        
        # Format compliance rate
        if label == "Compliance Rate":
            cell = ws.cell(row=row_idx, column=2)
            cell.font = Font(bold=True, size=14)
            
            if compliance_rate >= 90:
                cell.font = Font(bold=True, size=14, color="006100")
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif compliance_rate >= 70:
                cell.font = Font(bold=True, size=14, color="9C5700")
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                cell.font = Font(bold=True, size=14, color="9C0006")
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    # Add compliance assessment
    ws['A13'] = "Assessment:"
    ws['A13'].font = Font(bold=True, size=12)
    
    if compliance_rate >= 90:
        assessment = "✅ Excellent - HARA demonstrates strong ISO 26262 compliance"
        color = "006100"
    elif compliance_rate >= 70:
        assessment = "⚠️ Good - HARA meets most requirements with some improvements needed"
        color = "9C5700"
    elif compliance_rate >= 50:
        assessment = "⚠️ Fair - HARA requires significant improvements"
        color = "9C5700"
    else:
        assessment = "❌ Poor - HARA has major compliance issues requiring substantial rework"
        color = "9C0006"
    
    ws['A14'] = assessment
    ws['A14'].font = Font(size=11, color=color, italic=True)
    ws.merge_cells('A14:B14')


def create_category_breakdown_sheet(wb, reviews):
    """Create category-wise breakdown sheet."""
    ws = wb.create_sheet("Category Breakdown")
    
    # Group by category
    categories = {}
    for review in reviews:
        cat = review.get('category', 'Uncategorized')
        if cat not in categories:
            categories[cat] = {'pass': 0, 'fail': 0, 'partial': 0, 'na': 0, 'total': 0}
        
        categories[cat]['total'] += 1
        status = review.get('status', '').lower()
        
        if 'pass' in status and 'partial' not in status:
            categories[cat]['pass'] += 1
        elif 'fail' in status and 'partial' not in status:
            categories[cat]['fail'] += 1
        elif 'partial' in status:
            categories[cat]['partial'] += 1
        elif 'not applicable' in status or 'n/a' in status:
            categories[cat]['na'] += 1
    
    # Headers
    headers = ["Category", "Total", "Pass", "Fail", "Partial", "N/A", "Compliance %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="00467F", end_color="00467F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write category data
    for row_idx, (category, stats) in enumerate(categories.items(), 2):
        applicable = stats['total'] - stats['na']
        compliance = (stats['pass'] / applicable * 100) if applicable > 0 else 0
        
        data = [
            category,
            stats['total'],
            stats['pass'],
            stats['fail'],
            stats['partial'],
            stats['na'],
            f"{compliance:.1f}%"
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical="center")
            
            # Color code compliance percentage
            if col_idx == 7:  # Compliance % column
                if compliance >= 90:
                    cell.font = Font(bold=True, color="006100")
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif compliance >= 70:
                    cell.font = Font(bold=True, color="9C5700")
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.font = Font(bold=True, color="9C0006")
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 12
    
    # Freeze header
    ws.freeze_panes = "A2"