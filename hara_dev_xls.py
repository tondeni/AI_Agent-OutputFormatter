# hara_dev_xls.py - HARA Table Excel formatter
# Place this file in: AI_Agent-OutputFormatter plugin folder

from datetime import datetime
from cat.log import log
import re

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    log.warning("openpyxl not available - HARA Excel export will be disabled")


def parse_hara_table(content):
    """
    Parse HARA table from markdown format into structured data.
    Handles both 10-column (legacy) and 12-column (with Safe State & FTTI) formats.
    
    Args:
        content (str): Markdown content containing HARA table
        
    Returns:
        list: List of HARA entries as dictionaries
    """
    hara_entries = []
    
    # Find table rows (lines starting with |)
    lines = content.split('\n')
    in_table = False
    headers = []
    
    for line in lines:
        line = line.strip()
        if not line.startswith('|'):
            continue
            
        # Split by | and clean
        cells = [cell.strip() for cell in line.split('|')[1:-1]]
        
        if not cells:
            continue
        
        # Check if this is a header row (contains "Hazard ID" or similar)
        if any(keyword in cells[0] for keyword in ['Hazard ID', 'ID', 'HAZ-']):
            if 'Hazard ID' in cells[0] or 'ID' in cells[0]:
                headers = cells
                in_table = True
                log.info(f"Found HARA table with {len(headers)} columns")
                continue
        
        # Check if this is a separator row (with dashes)
        if all('-' in cell for cell in cells[:3]):
            continue
        
        # This is a data row
        if in_table and len(cells) >= 10:
            # Support both 10-column and 12-column formats
            entry = {
                'hazard_id': cells[0],
                'function': cells[1],
                'malfunction': cells[2],
                'hazard': cells[3],
                'situation': cells[4],
                'severity': cells[5],
                'exposure': cells[6],
                'controllability': cells[7],
                'asil': cells[8],
                'safety_goal': cells[9],
                'safe_state': cells[10] if len(cells) > 10 else 'N/A',
                'ftti': cells[11] if len(cells) > 11 else 'N/A'
            }
            hara_entries.append(entry)
    
    log.info(f"Parsed {len(hara_entries)} HARA entries from table")
    return hara_entries


def create_hara_excel(hara_entries, system_name, timestamp):
    """
    Create an Excel file with HARA table including Safe State and FTTI columns.
    
    Args:
        hara_entries (list): List of HARA entry dictionaries
        system_name (str): System name
        timestamp (str): Timestamp string
        
    Returns:
        Workbook: openpyxl Workbook object or None
    """
    if not EXCEL_AVAILABLE:
        log.warning("openpyxl not available - cannot create Excel file")
        return None
    
    log.info(f"Creating HARA Excel with {len(hara_entries)} entries")
    
    wb = openpyxl.Workbook()
    
    # Create sheets in order
    create_hara_summary_sheet(wb, hara_entries, system_name)  # First sheet
    create_safety_goals_summary(wb, hara_entries)              # Second sheet
    create_hara_table_sheet(wb, hara_entries, system_name)     # Third sheet (will be active)
    
    log.info("HARA Excel created successfully")
    return wb


def create_hara_table_sheet(wb, hara_entries, system_name):
    """Create the main HARA table sheet with all columns including Safe State and FTTI."""
    
    # Get the active sheet or create new one
    if "Sheet" in [ws.title for ws in wb.worksheets]:
        ws = wb["Sheet"]
        ws.title = "HARA Table"
    else:
        ws = wb.create_sheet("HARA Table")
    
    # Title rows
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = f"HARA Table: {system_name}"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('A2:L2')
    date_cell = ws['A2']
    date_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    date_cell.font = Font(italic=True)
    date_cell.alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A3:L3')
    standard_cell = ws['A3']
    standard_cell.value = "ISO 26262-3:2018 - Clause 6"
    standard_cell.font = Font(italic=True, size=10)
    standard_cell.alignment = Alignment(horizontal="center")
    
    # Headers (row 5) - Now with 12 columns
    headers = [
        'Hazard ID', 
        'Function', 
        'Malfunctioning Behavior', 
        'Hazardous Event',
        'Operational Situation',
        'Severity (S)',
        'Exposure (E)',
        'Controllability (C)',
        'ASIL',
        'Safety Goal',
        'Safe State',
        'FTTI'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Data rows (starting from row 6)
    for row_idx, entry in enumerate(hara_entries, 6):
        data = [
            entry['hazard_id'],
            entry['function'],
            entry['malfunction'],
            entry['hazard'],
            entry['situation'],
            entry['severity'],
            entry['exposure'],
            entry['controllability'],
            entry['asil'],
            entry['safety_goal'],
            entry['safe_state'],
            entry['ftti']
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Color code ASIL column (column 9)
            if col == 9:
                apply_asil_formatting(cell, value)
            
            # Color code S/E/C columns (columns 6, 7, 8)
            if col in [6, 7, 8]:
                apply_sec_formatting(cell, value)
    
    # Adjust column widths
    column_widths = {
        'A': 12,  # Hazard ID
        'B': 20,  # Function
        'C': 25,  # Malfunction
        'D': 25,  # Hazard
        'E': 20,  # Situation
        'F': 10,  # S
        'G': 10,  # E
        'H': 15,  # C
        'I': 8,   # ASIL
        'J': 35,  # Safety Goal
        'K': 25,  # Safe State
        'L': 10   # FTTI
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Adjust row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[5].height = 40
    
    # Freeze panes (freeze header rows)
    ws.freeze_panes = 'A6'


def create_safety_goals_summary(wb, hara_entries):
    """
    Create Safety Goals summary sheet with unique goals and their max ASIL.
    
    Args:
        wb: openpyxl Workbook object
        hara_entries (list): List of HARA entry dictionaries
    """
    summary_ws = wb.create_sheet("Safety Goals Summary", 1)  # Insert as second sheet
    
    # Extract unique safety goals with their ASIL levels
    safety_goals_dict = {}
    
    for entry in hara_entries:
        sg = entry['safety_goal'].strip()
        asil = entry['asil'].strip()
        
        if sg not in safety_goals_dict:
            safety_goals_dict[sg] = []
        
        safety_goals_dict[sg].append(asil)
    
    # Determine max ASIL for each goal
    asil_priority = {'QM': 0, 'A': 1, 'B': 2, 'C': 3, 'D': 4}
    
    safety_goals_summary = []
    for sg, asil_list in safety_goals_dict.items():
        # Find max ASIL
        max_asil = 'QM'
        max_priority = 0
        
        for asil in asil_list:
            # Extract ASIL letter (handle formats like "ASIL D", "D", etc.)
            asil_clean = asil.upper().replace('ASIL', '').strip()
            if asil_clean in asil_priority:
                if asil_priority[asil_clean] > max_priority:
                    max_priority = asil_priority[asil_clean]
                    max_asil = asil_clean
        
        safety_goals_summary.append({
            'goal': sg,
            'max_asil': max_asil,
            'occurrences': len(asil_list)
        })
    
    # Sort by ASIL (highest first)
    safety_goals_summary.sort(key=lambda x: asil_priority[x['max_asil']], reverse=True)
    
    # Title
    summary_ws.merge_cells('A1:C1')
    title_cell = summary_ws['A1']
    title_cell.value = "Safety Goals Summary"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    summary_ws.row_dimensions[1].height = 25
    
    # Description
    summary_ws.merge_cells('A2:C2')
    desc_cell = summary_ws['A2']
    desc_cell.value = "Unique Safety Goals with Maximum ASIL Level"
    desc_cell.font = Font(italic=True, size=10)
    desc_cell.alignment = Alignment(horizontal="center")
    
    # Create header (row 4)
    headers = ["Safety Goal", "Maximum ASIL", "Occurrences"]
    
    for col, header in enumerate(headers, 1):
        cell = summary_ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Add data rows (starting from row 5)
    for row_idx, sg_data in enumerate(safety_goals_summary, 5):
        summary_ws.cell(row=row_idx, column=1, value=sg_data['goal'])
        summary_ws.cell(row=row_idx, column=2, value=sg_data['max_asil'])
        summary_ws.cell(row=row_idx, column=3, value=sg_data['occurrences'])
        
        for col in range(1, 4):
            cell = summary_ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Color code ASIL column
            if col == 2:
                apply_asil_formatting(cell, sg_data['max_asil'])
    
    # Set column widths
    summary_ws.column_dimensions['A'].width = 50
    summary_ws.column_dimensions['B'].width = 15
    summary_ws.column_dimensions['C'].width = 12
    
    # Freeze header
    summary_ws.freeze_panes = 'A5'


def create_hara_summary_sheet(wb, hara_entries, system_name):
    """Create a summary statistics sheet for HARA."""
    
    # Remove default sheet if it exists and create Summary as first sheet
    if "Sheet" in [ws.title for ws in wb.worksheets]:
        wb.remove(wb["Sheet"])
    
    summary_ws = wb.create_sheet("Summary", 0)  # Insert as first sheet
    
    # Title
    summary_ws.merge_cells('A1:B1')
    title_cell = summary_ws['A1']
    title_cell.value = f"HARA Summary: {system_name}"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    summary_ws.row_dimensions[1].height = 25
    
    # Count by ASIL
    asil_counts = {'QM': 0, 'A': 0, 'B': 0, 'C': 0, 'D': 0}
    for entry in hara_entries:
        asil = entry['asil'].upper()
        for key in asil_counts.keys():
            if key in asil:
                asil_counts[key] += 1
                break
    
    # Count by Severity
    severity_counts = {'S0': 0, 'S1': 0, 'S2': 0, 'S3': 0}
    for entry in hara_entries:
        severity = entry['severity'].upper()
        for key in severity_counts.keys():
            if key in severity:
                severity_counts[key] += 1
                break
    
    # Count by Exposure
    exposure_counts = {'E0': 0, 'E1': 0, 'E2': 0, 'E3': 0}
    for entry in hara_entries:
        exposure = entry['exposure'].upper()
        for key in exposure_counts.keys():
            if key in exposure:
                exposure_counts[key] += 1
                break
    
    # Count by Controllability
    controllability_counts = {'C0': 0, 'C1': 0, 'C2': 0, 'C3': 0}
    for entry in hara_entries:
        controllability = entry['controllability'].upper()
        for key in controllability_counts.keys():
            if key in controllability:
                controllability_counts[key] += 1
                break
    
    # Summary data
    summary_data = [
        ["", ""],
        ["Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ["Total Hazards:", len(hara_entries)],
        ["", ""],
        ["ASIL Distribution", "Count"],
        ["ASIL D (Highest):", asil_counts['D']],
        ["ASIL C:", asil_counts['C']],
        ["ASIL B:", asil_counts['B']],
        ["ASIL A:", asil_counts['A']],
        ["QM (No ASIL):", asil_counts['QM']],
        ["", ""],
        ["Severity Distribution", "Count"],
        ["S3 (Life-threatening/Fatal):", severity_counts['S3']],
        ["S2 (Severe injuries):", severity_counts['S2']],
        ["S1 (Light/moderate injuries):", severity_counts['S1']],
        ["S0 (No injuries):", severity_counts['S0']],
        ["", ""],
        ["Exposure Distribution", "Count"],
        ["E3 (Medium-High probability):", exposure_counts['E3']],
        ["E2 (Low probability):", exposure_counts['E2']],
        ["E1 (Very low probability):", exposure_counts['E1']],
        ["E0 (Incredible):", exposure_counts['E0']],
        ["", ""],
        ["Controllability Distribution", "Count"],
        ["C3 (Difficult/Uncontrollable):", controllability_counts['C3']],
        ["C2 (Normally controllable):", controllability_counts['C2']],
        ["C1 (Simply controllable):", controllability_counts['C1']],
        ["C0 (Controllable in general):", controllability_counts['C0']],
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, 3):
        summary_ws.cell(row=row_idx, column=1, value=label)
        summary_ws.cell(row=row_idx, column=2, value=value)
        
        # Style headers
        if label in ["ASIL Distribution", "Severity Distribution", 
                     "Exposure Distribution", "Controllability Distribution"]:
            summary_ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
        elif label and ":" in label and value != "":
            summary_ws.cell(row=row_idx, column=1).font = Font(bold=True)
            
            # Color code ASIL rows
            if "ASIL D" in label:
                summary_ws.cell(row=row_idx, column=1).font = Font(bold=True, color="9C0006")
                summary_ws.cell(row=row_idx, column=2).font = Font(bold=True, color="9C0006")
            elif "ASIL C" in label:
                summary_ws.cell(row=row_idx, column=1).font = Font(bold=True, color="9C6500")
                summary_ws.cell(row=row_idx, column=2).font = Font(bold=True, color="9C6500")
    
    summary_ws.column_dimensions['A'].width = 35
    summary_ws.column_dimensions['B'].width = 15
    
    # Add compliance note
    note_row = len(summary_data) + 5
    summary_ws.merge_cells(f'A{note_row}:B{note_row}')
    note_cell = summary_ws.cell(row=note_row, column=1)
    note_cell.value = "ISO 26262-3:2018 Compliance"
    note_cell.font = Font(bold=True, size=12)
    note_cell.alignment = Alignment(horizontal="center")
    
    compliance_items = [
        "✓ Clause 6.4.3 - Hazard identification (HAZOP)",
        "✓ Clause 6.4.4 - Classification of hazardous events (S, E, C)",
        "✓ Clause 6.4.5 - ASIL determination",
        "✓ Clause 6.4.6 - Safety goal determination",
    ]
    
    for idx, item in enumerate(compliance_items, 1):
        summary_ws.cell(row=note_row + idx, column=1, value=item)
        summary_ws.cell(row=note_row + idx, column=1).font = Font(color="006100")


def apply_asil_formatting(cell, asil_value):
    """
    Apply color formatting based on ASIL value.
    
    Args:
        cell: openpyxl Cell object
        asil_value (str): ASIL level
    """
    asil_clean = str(asil_value).upper().replace('ASIL', '').strip()
    
    if asil_clean == 'D':
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cell.font = Font(bold=True, color="9C0006")
    elif asil_clean == 'C':
        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        cell.font = Font(bold=True, color="9C6500")
    elif asil_clean == 'B':
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cell.font = Font(bold=True, color="006100")
    elif asil_clean == 'A':
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = Font(bold=True)
    elif asil_clean == 'QM':
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")


def apply_sec_formatting(cell, value):
    """
    Apply color formatting for S/E/C values.
    
    Args:
        cell: openpyxl Cell object
        value (str): S/E/C value
    """
    val_clean = str(value).upper().strip()
    
    if val_clean in ['S3', 'E3', 'C3']:
        cell.fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
        cell.font = Font(bold=True)
    elif val_clean in ['S2', 'E2', 'C2']:
        cell.fill = PatternFill(start_color="FFDB99", end_color="FFDB99", fill_type="solid")
    elif val_clean in ['S1', 'E1', 'C1']:
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    elif val_clean in ['S0', 'E0', 'C0']:
        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")