# ==============================================================================
# AI_Agent-OutputFormatter/code/generators/fsc_excel_generator.py  
# Self-contained FSC Excel generator with validation and preparation
# ==============================================================================

from typing import List, Dict, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class FSCExcelGenerator:
    """
    Self-contained generator for FSC Excel workbooks.
    
    Handles:
    - Data validation
    - Statistics calculation
    - Multi-sheet organization
    - Excel file creation
    """
    
    def __init__(self):
        """Initialize Excel generator"""
        self.header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        self.header_font = Font(bold=True, size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def validate_data(
        self,
        goals_data: List[Dict],
        fsrs_data: List[Dict]
    ) -> Tuple[bool, List[str], List[str]]:
        """
        Validate FSC data for Excel generation.
        
        Returns:
            Tuple of (is_valid, warnings, errors)
        """
        warnings = []
        errors = []
        
        if not goals_data:
            errors.append("No safety goals available")
        
        if not fsrs_data:
            errors.append("No FSRs available")
        
        if errors:
            return False, warnings, errors
        
        # Check for duplicate IDs
        fsr_ids = [fsr.get('id', fsr.get('fsr_id', '')) for fsr in fsrs_data]
        if len(fsr_ids) != len(set(fsr_ids)):
            warnings.append("Duplicate FSR IDs detected")
        
        return True, warnings, errors
    
    def calculate_statistics(
        self,
        goals_data: List[Dict],
        fsrs_data: List[Dict]
    ) -> Dict:
        """Calculate Excel-specific statistics"""
        
        # ASIL distribution
        asil_dist = {}
        for fsr in fsrs_data:
            asil = fsr.get('asil', 'QM')
            asil_dist[asil] = asil_dist.get(asil, 0) + 1
        
        # Type distribution
        type_dist = {}
        for fsr in fsrs_data:
            fsr_type = fsr.get('type', 'other')
            type_dist[fsr_type] = type_dist.get(fsr_type, 0) + 1
        
        # Allocation status
        allocated = sum(1 for fsr in fsrs_data if fsr.get('allocated_to'))
        
        return {
            'total_goals': len(goals_data),
            'total_fsrs': len(fsrs_data),
            'asil_distribution': asil_dist,
            'type_distribution': type_dist,
            'allocated_fsrs': allocated,
            'unallocated_fsrs': len(fsrs_data) - allocated
        }
    
    def generate(
        self,
        system_name: str,
        goals_data: List[Dict],
        fsrs_data: List[Dict],
        strategies_data: Optional[Dict] = None,
        allocation_data: Optional[Dict] = None
    ) -> Workbook:
        """
        Generate complete FSC Excel workbook.
        
        Args:
            system_name: Name of the system/item
            goals_data: Safety goals from HARA
            fsrs_data: Functional Safety Requirements
            strategies_data: Safety strategies (optional)
            allocation_data: Allocation matrix (optional)
            
        Returns:
            Workbook object ready to save
        """
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets
        self._create_safety_goals_sheet(wb, goals_data)
        self._create_fsrs_sheet(wb, fsrs_data)
        self._create_allocation_sheet(wb, fsrs_data, allocation_data or {})
        self._create_traceability_sheet(wb, goals_data, fsrs_data)
        self._create_statistics_sheet(wb, goals_data, fsrs_data)
        
        return wb
    
    def _create_safety_goals_sheet(self, wb: Workbook, goals_data: List[Dict]):
        """Create Safety Goals sheet"""
        ws = wb.create_sheet("Safety Goals")
        
        # Headers
        headers = ["SG-ID", "Safety Goal", "ASIL", "Safe State", "FTTI (ms)", "FSR Count"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data
        for row, goal in enumerate(goals_data, 2):
            ws.cell(row=row, column=1, value=goal.get('sg_id', goal.get('id', '')))
            ws.cell(row=row, column=2, value=goal.get('statement', goal.get('goal', '')))
            ws.cell(row=row, column=3, value=goal.get('asil', 'QM'))
            ws.cell(row=row, column=4, value=goal.get('safe_state', 'TBD'))
            ws.cell(row=row, column=5, value=goal.get('ftti_ms', goal.get('ftti', 'TBD')))
            ws.cell(row=row, column=6, value=0)  # Placeholder for FSR count
            
            # Apply borders
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.border
        
        # Auto-size columns
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        ws.column_dimensions['B'].width = 50  # Safety Goal statement
    
    def _create_fsrs_sheet(self, wb: Workbook, fsrs_data: List[Dict]):
        """Create FSRs sheet"""
        ws = wb.create_sheet("Functional Safety Requirements")
        
        # Headers
        headers = ["FSR-ID", "Description", "Type", "ASIL", "Parent SG", "Allocated To", "FHTI (ms)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data
        for row, fsr in enumerate(fsrs_data, 2):
            ws.cell(row=row, column=1, value=fsr.get('id', fsr.get('fsr_id', '')))
            ws.cell(row=row, column=2, value=fsr.get('description', fsr.get('requirement', '')))
            ws.cell(row=row, column=3, value=fsr.get('type', 'other'))
            ws.cell(row=row, column=4, value=fsr.get('asil', 'QM'))
            
            # Parent SG
            parent_sg = fsr.get('parent_safety_goal', fsr.get('parent_sg', []))
            if isinstance(parent_sg, list):
                ws.cell(row=row, column=5, value=', '.join(parent_sg))
            else:
                ws.cell(row=row, column=5, value=str(parent_sg))
            
            ws.cell(row=row, column=6, value=fsr.get('allocated_to', 'TBD'))
            ws.cell(row=row, column=7, value=fsr.get('fhti_ms', fsr.get('ftti_ms', 'TBD')))
            
            # Apply borders
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.border
        
        # Auto-size columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        ws.column_dimensions['B'].width = 60  # Description
        ws.column_dimensions['F'].width = 25  # Allocated To
    
    def _create_allocation_sheet(self, wb: Workbook, fsrs_data: List[Dict], allocation_data: Dict):
        """Create Allocation Matrix sheet"""
        ws = wb.create_sheet("Allocation Matrix")
        
        # Group FSRs by component
        by_component = {}
        for fsr in fsrs_data:
            component = fsr.get('allocated_to', 'Unallocated')
            if component not in by_component:
                by_component[component] = []
            by_component[component].append(fsr)
        
        # Headers
        headers = ["Component", "FSR Count", "ASIL Levels", "FSR IDs"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data
        row = 2
        for component, fsrs in sorted(by_component.items()):
            ws.cell(row=row, column=1, value=component)
            ws.cell(row=row, column=2, value=len(fsrs))
            
            # ASIL levels
            asils = set(fsr.get('asil', 'QM') for fsr in fsrs)
            ws.cell(row=row, column=3, value=', '.join(sorted(asils)))
            
            # FSR IDs
            fsr_ids = [fsr.get('id', fsr.get('fsr_id', '')) for fsr in fsrs]
            ws.cell(row=row, column=4, value=', '.join(fsr_ids[:10]) + ('...' if len(fsr_ids) > 10 else ''))
            
            # Apply borders
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50
    
    def _create_traceability_sheet(self, wb: Workbook, goals_data: List[Dict], fsrs_data: List[Dict]):
        """Create Traceability Matrix sheet"""
        ws = wb.create_sheet("Traceability")
        
        # Headers
        headers = ["Safety Goal ID", "Safety Goal", "ASIL", "FSR IDs", "FSR Count"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data
        for row, goal in enumerate(goals_data, 2):
            goal_id = goal.get('sg_id', goal.get('id', ''))
            
            ws.cell(row=row, column=1, value=goal_id)
            ws.cell(row=row, column=2, value=goal.get('statement', goal.get('goal', '')))
            ws.cell(row=row, column=3, value=goal.get('asil', 'QM'))
            
            # Find related FSRs
            related_fsrs = []
            for fsr in fsrs_data:
                parent_sg = fsr.get('parent_safety_goal', fsr.get('parent_sg', []))
                if isinstance(parent_sg, list):
                    if goal_id in parent_sg:
                        related_fsrs.append(fsr.get('id', fsr.get('fsr_id', '')))
                elif parent_sg == goal_id:
                    related_fsrs.append(fsr.get('id', fsr.get('fsr_id', '')))
            
            ws.cell(row=row, column=4, value=', '.join(related_fsrs[:10]) + ('...' if len(related_fsrs) > 10 else ''))
            ws.cell(row=row, column=5, value=len(related_fsrs))
            
            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.border
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 12
    
    def _create_statistics_sheet(self, wb: Workbook, goals_data: List[Dict], fsrs_data: List[Dict]):
        """Create Statistics sheet"""
        ws = wb.create_sheet("Statistics")
        
        stats = self.calculate_statistics(goals_data, fsrs_data)
        
        row = 1
        
        # Title
        ws.cell(row=row, column=1, value="FSC Statistics Summary")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2
        
        # Overall counts
        ws.cell(row=row, column=1, value="Category")
        ws.cell(row=row, column=2, value="Count")
        ws.cell(row=row, column=1).font = self.header_font
        ws.cell(row=row, column=2).font = self.header_font
        ws.cell(row=row, column=1).fill = self.header_fill
        ws.cell(row=row, column=2).fill = self.header_fill
        row += 1
        
        ws.cell(row=row, column=1, value="Total Safety Goals")
        ws.cell(row=row, column=2, value=stats['total_goals'])
        row += 1
        
        ws.cell(row=row, column=1, value="Total FSRs")
        ws.cell(row=row, column=2, value=stats['total_fsrs'])
        row += 2
        
        # ASIL Distribution
        ws.cell(row=row, column=1, value="ASIL Distribution")
        ws.cell(row=row, column=1).font = self.header_font
        row += 1
        
        for asil, count in sorted(stats['asil_distribution'].items()):
            ws.cell(row=row, column=1, value=f"ASIL {asil}")
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        row += 1
        
        # Type Distribution
        ws.cell(row=row, column=1, value="FSR Type Distribution")
        ws.cell(row=row, column=1).font = self.header_font
        row += 1
        
        for fsr_type, count in sorted(stats['type_distribution'].items()):
            ws.cell(row=row, column=1, value=fsr_type.capitalize())
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        row += 1
        
        # Allocation Status
        ws.cell(row=row, column=1, value="Allocation Status")
        ws.cell(row=row, column=1).font = self.header_font
        row += 1
        
        ws.cell(row=row, column=1, value="Allocated FSRs")
        ws.cell(row=row, column=2, value=stats['allocated_fsrs'])
        row += 1
        
        ws.cell(row=row, column=1, value="Unallocated FSRs")
        ws.cell(row=row, column=2, value=stats['unallocated_fsrs'])
        row += 1
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15


# Convenience function for direct use
def generate_fsc_excel(
    system_name: str,
    goals_data: List[Dict],
    fsrs_data: List[Dict],
    strategies_data: Optional[Dict] = None,
    allocation_data: Optional[Dict] = None
) -> Workbook:
    """
    Generate FSC Excel workbook.
    
    Convenience function that creates generator and generates workbook.
    """
    generator = FSCExcelGenerator()
    return generator.generate(
        system_name, goals_data, fsrs_data, strategies_data, allocation_data
    )