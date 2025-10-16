# ==============================================================================
# fsr_excel_generator.py
# Excel file generation for FSC data
# Place in: AI_Agent-OutputFormatter/generators/fsr_excel_generator.py
# ==============================================================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_fsr_excel(fsrs_data, goals_data, system_name):
    """
    Generate Excel workbook with FSC data.
    
    Args:
        fsrs_data: List of FSR dictionaries from working memory
        goals_data: List of Safety Goal dictionaries from working memory
        system_name: Name of the system
    
    Returns:
        openpyxl Workbook object
    """
    
    wb = Workbook()
    
    # Create worksheets
    ws_fsrs = wb.active
    ws_fsrs.title = "FSRs"
    
    ws_allocation = wb.create_sheet("Allocation Matrix")
    ws_traceability = wb.create_sheet("Traceability")
    ws_statistics = wb.create_sheet("Statistics")
    
    # Populate worksheets
    _create_fsr_sheet(ws_fsrs, fsrs_data, system_name)
    _create_allocation_sheet(ws_allocation, fsrs_data, system_name)
    _create_traceability_sheet(ws_traceability, fsrs_data, goals_data, system_name)
    _create_statistics_sheet(ws_statistics, fsrs_data, system_name)
    
    return wb


def _create_fsr_sheet(ws, fsrs_data, system_name):
    """Create FSR listing sheet."""
    
    # Title
    ws['A1'] = f"Functional Safety Requirements - {system_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')
    
    # Headers
    headers = [
        "FSR ID", "Type", "ASIL", "Description",
        "Safety Goal", "Allocated To", "Verification", "Status"
    ]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for row_idx, fsr in enumerate(fsrs_data, 4):
        ws.cell(row=row_idx, column=1).value = fsr.get('id', '')
        ws.cell(row=row_idx, column=2).value = fsr.get('type', '')
        ws.cell(row=row_idx, column=3).value = fsr.get('asil', '')
        ws.cell(row=row_idx, column=4).value = fsr.get('description', '')
        ws.cell(row=row_idx, column=5).value = fsr.get('safety_goal_id', '')
        ws.cell(row=row_idx, column=6).value = fsr.get('allocated_to', 'Not allocated')
        ws.cell(row=row_idx, column=7).value = fsr.get('verification_criteria', 'TBD')
        
        # Status based on allocation
        status = "Allocated" if fsr.get('allocated_to') else "Pending Allocation"
        ws.cell(row=row_idx, column=8).value = status
    
    # Auto-adjust column widths
    for col in range(1, 9):
        max_length = 0
        column_letter = get_column_letter(col)
        
        for cell in ws[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header
    ws.freeze_panes = 'A4'


def _create_allocation_sheet(ws, fsrs_data, system_name):
    """Create allocation matrix sheet."""
    
    # Title
    ws['A1'] = f"FSR Allocation Matrix - {system_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:E1')
    
    # Group FSRs by component
    by_component = {}
    unallocated = []
    
    for fsr in fsrs_data:
        component = fsr.get('allocated_to')
        if component and component not in ['TBD', 'NOT ALLOCATED', 'N/A', '']:
            if component not in by_component:
                by_component[component] = []
            by_component[component].append(fsr)
        else:
            unallocated.append(fsr)
    
    # Headers
    headers = ["Component", "FSR Count", "ASIL Levels", "FSR IDs", "Types"]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    row_idx = 4
    
    for component, fsrs in sorted(by_component.items()):
        ws.cell(row=row_idx, column=1).value = component
        ws.cell(row=row_idx, column=2).value = len(fsrs)
        
        # ASIL levels
        asil_levels = ', '.join(sorted(set(f.get('asil', 'QM') for f in fsrs), reverse=True))
        ws.cell(row=row_idx, column=3).value = asil_levels
        
        # FSR IDs
        fsr_ids = ', '.join(f.get('id', '') for f in fsrs[:5])
        if len(fsrs) > 5:
            fsr_ids += f" ... (+{len(fsrs) - 5} more)"
        ws.cell(row=row_idx, column=4).value = fsr_ids
        
        # Types
        types = ', '.join(set(f.get('type', '') for f in fsrs))
        ws.cell(row=row_idx, column=5).value = types
        
        row_idx += 1
    
    # Unallocated FSRs
    if unallocated:
        row_idx += 1
        cell = ws.cell(row=row_idx, column=1)
        cell.value = "⚠️ UNALLOCATED"
        cell.font = Font(bold=True, color="FF0000")
        
        ws.cell(row=row_idx, column=2).value = len(unallocated)
        
        fsr_ids = ', '.join(f.get('id', '') for f in unallocated[:5])
        if len(unallocated) > 5:
            fsr_ids += f" ... (+{len(unallocated) - 5} more)"
        ws.cell(row=row_idx, column=4).value = fsr_ids
    
    # Auto-adjust widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 25


def _create_traceability_sheet(ws, fsrs_data, goals_data, system_name):
    """Create traceability matrix with enhanced information from safety goals."""
    
    ws['A1'] = f"Traceability Matrix - {system_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:F1')
    
    # Headers
    headers = ["Safety Goal ID", "Safety Goal Description", "FSR ID", "FSR Type", "ASIL", "Allocated To"]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Create a lookup dictionary for safety goals
    goals_dict = {}
    if goals_data:
        for goal in goals_data:
            goal_id = goal.get('id', goal.get('goal_id', ''))
            if goal_id:
                goals_dict[goal_id] = goal
    
    # Group by safety goal
    by_goal = {}
    for fsr in fsrs_data:
        sg_id = fsr.get('safety_goal_id', 'Unknown')
        if sg_id not in by_goal:
            by_goal[sg_id] = []
        by_goal[sg_id].append(fsr)
    
    # Data rows
    row_idx = 4
    
    for sg_id, fsrs in sorted(by_goal.items()):
        # Get safety goal description if available
        sg_description = ""
        if sg_id in goals_dict:
            sg_description = goals_dict[sg_id].get('description', goals_dict[sg_id].get('goal', ''))[:100]
        
        for idx, fsr in enumerate(fsrs):
            # Only show safety goal description on first FSR of each goal
            if idx == 0:
                ws.cell(row=row_idx, column=1).value = sg_id
                ws.cell(row=row_idx, column=2).value = sg_description
                ws.cell(row=row_idx, column=1).font = Font(bold=True)
            else:
                ws.cell(row=row_idx, column=1).value = ""
                ws.cell(row=row_idx, column=2).value = ""
            
            ws.cell(row=row_idx, column=3).value = fsr.get('id', '')
            ws.cell(row=row_idx, column=4).value = fsr.get('type', '')
            ws.cell(row=row_idx, column=5).value = fsr.get('asil', '')
            ws.cell(row=row_idx, column=6).value = fsr.get('allocated_to', 'Not allocated')
            row_idx += 1
    
    # Auto-adjust widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 25


def _create_statistics_sheet(ws, fsrs_data, system_name):
    """Create statistics summary."""
    
    ws['A1'] = f"FSC Statistics - {system_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:C1')
    
    # Calculate statistics
    total = len(fsrs_data)
    allocated = sum(1 for f in fsrs_data 
                   if f.get('allocated_to') and f.get('allocated_to') not in ['TBD', 'NOT ALLOCATED', 'N/A', ''])
    
    by_asil = {}
    by_type = {}
    by_component = {}
    
    for fsr in fsrs_data:
        asil = fsr.get('asil', 'QM')
        by_asil[asil] = by_asil.get(asil, 0) + 1
        
        fsr_type = fsr.get('type', 'Unknown')
        by_type[fsr_type] = by_type.get(fsr_type, 0) + 1
        
        component = fsr.get('allocated_to')
        if component and component not in ['TBD', 'NOT ALLOCATED', 'N/A', '']:
            by_component[component] = by_component.get(component, 0) + 1
    
    # Write statistics
    row = 3
    
    ws.cell(row=row, column=1).value = "Total FSRs"
    ws.cell(row=row, column=2).value = total
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    ws.cell(row=row, column=1).value = "Allocated"
    ws.cell(row=row, column=2).value = allocated
    ws.cell(row=row, column=3).value = f"{allocated/total*100:.1f}%" if total > 0 else "0%"
    row += 1
    
    ws.cell(row=row, column=1).value = "Unallocated"
    ws.cell(row=row, column=2).value = total - allocated
    ws.cell(row=row, column=3).value = f"{(total-allocated)/total*100:.1f}%" if total > 0 else "0%"
    row += 2
    
    # By ASIL
    ws.cell(row=row, column=1).value = "Distribution by ASIL"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    for asil in ['D', 'C', 'B', 'A', 'QM']:
        if asil in by_asil:
            ws.cell(row=row, column=1).value = f"ASIL {asil}"
            ws.cell(row=row, column=2).value = by_asil[asil]
            ws.cell(row=row, column=3).value = f"{by_asil[asil]/total*100:.1f}%"
            row += 1
    
    row += 1
    
    # By Type
    ws.cell(row=row, column=1).value = "Distribution by Type"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    for fsr_type, count in sorted(by_type.items()):
        ws.cell(row=row, column=1).value = fsr_type
        ws.cell(row=row, column=2).value = count
        ws.cell(row=row, column=3).value = f"{count/total*100:.1f}%"
        row += 1
    
    row += 1
    
    # By Component
    if by_component:
        ws.cell(row=row, column=1).value = "Distribution by Component"
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        for component, count in sorted(by_component.items(), key=lambda x: x[1], reverse=True):
            ws.cell(row=row, column=1).value = component
            ws.cell(row=row, column=2).value = count
            ws.cell(row=row, column=3).value = f"{count/total*100:.1f}%"
            row += 1
    
    # Auto-adjust
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15